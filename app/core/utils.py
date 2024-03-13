"""
Django Admin Actions Helpers.
"""

from django.contrib import admin
from django.contrib.admin import SimpleListFilter

from django.contrib.auth.admin import UserAdmin as BaseUserAdmin
from django.utils.translation import gettext_lazy as _
from rest_framework import status
from core import models
import openpyxl
from django.http import HttpResponse
from follower.serializers import FriendshipDetailSerializer
from comments.serializers import CommentSerializer
from posts.serializers import PostDetailSerializer
from django.contrib.auth import get_user_model

class LikesRangeFieldListFilter(admin.filters.FieldListFilter):
    template = 'html/filter_range_integer_input.html'
    parameter_name = 'likes'  # Update this with your actual parameter name

    def __init__(self, field, request, params, model, model_admin, field_path):
        super().__init__(field, request, params, model, model_admin, field_path)
        self.lookup_val_min = request.GET.get(self.parameter_name + '_min')
        self.lookup_val_max = request.GET.get(self.parameter_name + '_max')

    def queryset(self, request, queryset):
        if self.lookup_val_min and self.lookup_val_max:
            try:
                min_value = int(self.lookup_val_min)
                max_value = int(self.lookup_val_max)
                return queryset.filter(
                    likes__gte= min_value,
                    likes__lte = max_value)
            except ValueError:
                return queryset.none()
            
        return queryset

    def expected_parameters(self):
        return [self.parameter_name + '_min', self.parameter_name + '_max']

    def choices(self, changelist):
        return []

class OwnerTextFieldListFilter(admin.filters.FieldListFilter):
    template = 'html/filter_text_input.html'
    parameter_name = 'owner'  # Update this with your actual parameter name

    def __init__(self, field, request, params, model, model_admin, field_path):
        super().__init__(field, request, params, model, model_admin, field_path)
        self.owner = request.GET.get(self.parameter_name)

    def queryset(self, request, queryset):
        if self.owner:
            
            try:
                request_owner = str(self.owner)
                qs_for_name_email = get_user_model().objects.none()
                qs_for_id = queryset.none()
                try:
                    id = -1
                    try:
                        id = get_user_model().objects.filter(email = request_owner).first().id 
                    except:
                        id = get_user_model().objects.filter(name = request_owner).first().id
                    qs_for_name_email = queryset.filter(owner = id)
                except:
                    pass
                try:
                    qs_for_id = queryset.filter(
                        owner =request_owner
                    )
                except:
                    if qs_for_name_email.exists():
                        return qs_for_name_email
                    return qs_for_id
                return qs_for_id | qs_for_name_email
            except ValueError:
                pass
        return queryset

    def expected_parameters(self):
        return [self.parameter_name]

    def choices(self, changelist):
        return []


class FollowerTextFieldListFilter(admin.filters.FieldListFilter):
    template = 'html/filter_text_input.html'
    parameter_name = 'follower'  # Update this with your actual parameter name

    def __init__(self, field, request, params, model, model_admin, field_path):
        super().__init__(field, request, params, model, model_admin, field_path)
        self.user = request.GET.get(self.parameter_name)

    def queryset(self, request, queryset):
        if self.user:
            
            try:
                request_user = str(self.user)
                qs_for_name_email = get_user_model().objects.none()
                qs_for_id = queryset.none()
                try:
                    id = -1
                    try:
                        id = get_user_model().objects.filter(email = request_user).first().id 
                    except:
                        id = get_user_model().objects.filter(name = request_user).first().id
                    qs_for_name_email = queryset.filter(follower = id)
                except:
                    pass
                try:
                    qs_for_id = queryset.filter(
                        follower =request_user
                    )
                except:
                    if qs_for_name_email.exists():
                        return qs_for_name_email
                    return qs_for_id
                return qs_for_id | qs_for_name_email
            except ValueError:
                pass
        return queryset
    def expected_parameters(self):
        return [self.parameter_name]

    def choices(self, changelist):
        return []
    
class FollowingTextFieldListFilter(admin.filters.FieldListFilter):
    template = 'html/filter_text_input.html'
    parameter_name = 'following'  # Update this with your actual parameter name

    def __init__(self, field, request, params, model, model_admin, field_path):
        super().__init__(field, request, params, model, model_admin, field_path)
        self.user = request.GET.get(self.parameter_name)

    def queryset(self, request, queryset):
        if self.user:
            
            try:
                request_user = str(self.user)
                qs_for_name_email = get_user_model().objects.none()
                qs_for_id = queryset.none()
                try:
                    id = -1
                    try:
                        id = get_user_model().objects.filter(email = request_user).first().id 
                    except:
                        id = get_user_model().objects.filter(name = request_user).first().id
                    qs_for_name_email = queryset.filter(following = id)
                except:
                    pass
                try:
                    qs_for_id = queryset.filter(
                        following =request_user
                    )
                except:
                    if qs_for_name_email.exists():
                        return qs_for_name_email
                    return qs_for_id
                return qs_for_id | qs_for_name_email
            except ValueError:
                pass
        return queryset
    def expected_parameters(self):
        return [self.parameter_name]

    def choices(self, changelist):
        return []
class ContentFieldListFilter(admin.filters.FieldListFilter):
    template = 'html/filter_text_input.html'
    parameter_name = 'content'  # Update this with your actual parameter name

    def __init__(self, field, request, params, model, model_admin, field_path):
        super().__init__(field, request, params, model, model_admin, field_path)
        self.lookup_content = request.GET.get(self.parameter_name)

    def queryset(self, request, queryset):
        if self.lookup_content:
            try:
                content = str(self.lookup_content)
                return queryset.filter(
                    content= content
                    )
            except ValueError:
                return queryset.none()
        return queryset

    def expected_parameters(self):
        return [self.parameter_name]

    def choices(self, changelist):
        return []


class DateTimeRangeFilter(admin.filters.FieldListFilter):
    template = 'html/filter_datetime_range_input.html'
    parameter_name = 'createtime'
    def __init__(self, field, request, params, model, model_admin, field_path):
        super().__init__(field, request, params, model, model_admin, field_path)
        self.lookup_val_min = request.GET.get(self.parameter_name+ '_min')
        self.lookup_val_max = request.GET.get(self.parameter_name+ '_max')

    def queryset(self, request, queryset):
        if self.lookup_val_min and self.lookup_val_max:
            return queryset.filter(
                    **{self.field_path + '__range': (self.lookup_val_min, self.lookup_val_max)}
                )
        return queryset

    def _parse_date(self, date_str):
        try:
            return models.DateField().to_python(date_str)
        except:
            return None

    def expected_parameters(self):
        return [self.parameter_name + '_min', self.parameter_name + '_max']

    def choices(self, changelist):
        return []



class SinceDateTimeRangeFilter(admin.filters.FieldListFilter):
    template = 'html/filter_datetime_range_input.html'
    parameter_name = 'since'
    def __init__(self, field, request, params, model, model_admin, field_path):
        super().__init__(field, request, params, model, model_admin, field_path)
        self.lookup_val_min = request.GET.get(self.parameter_name+ '_min')
        self.lookup_val_max = request.GET.get(self.parameter_name+ '_max')

    def queryset(self, request, queryset):
        if self.lookup_val_min and self.lookup_val_max:
            return queryset.filter(
                    **{self.field_path + '__range': (self.lookup_val_min, self.lookup_val_max)}
                )
        return queryset

    def _parse_date(self, date_str):
        try:
            return models.DateField().to_python(date_str)
        except:
            return None

    def expected_parameters(self):
        return [self.parameter_name + '_min', self.parameter_name + '_max']

    def choices(self, changelist):
        return []

@admin.action(description="Download selected items as excel.")
def export_to_excel_friends(modeladmin, request, queryset):
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="exported_data.xlsx"'

    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = 'Exported Followship Data'
    # Write headers
    headers = ['id', 'follower', 'following', 'since','user_profile']
    worksheet.append(headers)

    # Write data rows
    for obj in FriendshipDetailSerializer(queryset,many=True).data:
        row = [
            obj['id'],
            obj['follower']['name'] if obj['follower'] else -1,
            obj['following'] if obj['following'] else -1,
            obj['since'],
            obj['follower']['user_profile']  if obj['follower'] else "",
        ]
        worksheet.append(row)

    workbook.save(response)
    return response

@admin.action(description="Download selected items as excel.")
def export_to_excel_post(modeladmin, request, queryset):
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="exported_data.xlsx"'

    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = 'Exported Post Data'
    # Write headers
    headers = ['id', 'owner', 'content', 'likes','comments', 'liked_users']
    worksheet.append(headers)

    # Write data rows
    for obj in PostDetailSerializer(queryset,many=True).data:
        row = [
            obj['id'],
            obj['owner']['name'] if obj['owner'] else "",
            obj['content'],
            obj['likes'],
            ', '.join([comment['content'] for comment in obj['comments']]) if obj['comments'] else '',
            ', '.join([user['name'] for user in obj['liked_users']]) if obj['liked_users'] else ''
        ]
        worksheet.append(row)

    workbook.save(response)
    return response


@admin.action(description="Download selected items as excel.")
def export_to_excel_comment(modeladmin, request, queryset):
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="exported_data.xlsx"'

    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = 'Exported Comment Data'
    # Write headers
    headers = ['id', 'related_post', 'owner', 'content', 'likes', 'liked_users']
    worksheet.append(headers)

    # Write data rows
    for obj in CommentSerializer(queryset,many=True).data:
        row = [
            obj['id'],
            obj['related_post'],
            obj['owner']['name'] if obj['owner'] else "",
            obj['content'],
            obj['likes'],
            ', '.join([user['name'] for user in obj['liked_users']]) if obj['liked_users'] else ''
        ]
        worksheet.append(row)

    workbook.save(response)
    return response