"""
Django Admin Costumization.
"""

from django.contrib import admin
from django.contrib.auth.admin import UserAdmin as BaseUserAdmin
from django.utils.translation import gettext_lazy as _
from rest_framework import status
from core import models
import openpyxl
from django.http import HttpResponse
from rest_framework.response import Response
import datetime
from . import utils

@admin.action(description="Download selected items as excel.")
def export_excel(modeladmin,request,queryset):
    if not queryset.exists():
        return HttpResponse(status = status.HTTP_400_BAD_REQUEST)
    
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="export.xlsx"'

    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = 'Exported User Data'

    # Write header row
    header = []
    first_item = queryset[0].__dict__.keys()

    i = 0
    for column in first_item:
        if(i != 0):
            header.append(column)
        i +=1
    for col_num, column_title in enumerate(header, 1):
        cell = worksheet.cell(row=1, column=col_num)
        cell.value = column_title

    # Write data rows

    for row_num, row in enumerate(queryset, 1):
        i = 0
        for col_num, cell_value in enumerate(row.__dict__.values(), 1):
            if i != 0:
                cell = worksheet.cell(row=row_num+1, column=col_num-1)
                #print(cell_value)
                if type(cell_value) == datetime.datetime:
                    cell.value = cell_value.replace(tzinfo=None)
                else:    
                    cell.value = cell_value
                    
            i+=1

    workbook.save(response)
    return response
class UserAdmin(BaseUserAdmin):
    """Define the admin pages for users."""
    ordering = ['id']
    list_display = ['email', 'name', 'last_login']
    list_filter = [('email', utils.OwnerTextFieldListFilter), ('name',utils.OwnerTextFieldListFilter), 'is_active', 'is_superuser', 'is_staff']
    search_fields = ['email','name']
    actions = [export_excel]
    fieldsets = (
        (None,{'fields': ('email', 'password', 'name',)}),
        (_('Permissions'),{'fields': ('is_active', 'is_superuser','is_staff',)}),
        (_('Important Dates'),{'fields': ('last_login',)}),
    )
    date_hierarchy = 'last_login'
    
    readonly_fields = ['last_login']
    add_fieldsets = (
        (None, {
            'classes': ('wide',), #Page look
            'fields': (
                'email',
                'password1',
                'password2',
                'name',
                'is_active',
                'is_staff',
                'is_superuser',
            ),
            }),
    )
class FollowerAdmin(admin.ModelAdmin):
    """Define the admin pages for followship."""
    ordering = ['id']
    list_display = ['id','following' ,'follower','since','profile_link']
    list_filter = [('follower', utils.FollowerTextFieldListFilter), ('following', utils.FollowingTextFieldListFilter), ('since',utils.SinceDateTimeRangeFilter)]
    search_fields = ['id','following__name','following__email','follower__name','follower__email']
    date_hierarchy = 'since'
    actions = [utils.export_to_excel_friends]
    fieldsets = (
        (None,{'fields': ('following','follower',)}),
        (_('Follower User Profile'),{'fields': ('profile_link',)}),
        (_('Important Dates'),{'fields': ('since',)}),
    )
    readonly_fields = ['since']

    add_fieldsets = (
        (None, {
            'classes': ('wide',), #Page look
            'fields': (
                'follower',
                'following',
                'since',
                'profile_link',
            ),
            }),
    )
class PostsAdmin(admin.ModelAdmin):
    """Define the admin pages for posts."""
    ordering = ['id']
    list_display = ['id','owner','content','likes','create_time']
    list_filter = [('owner',utils.OwnerTextFieldListFilter), ('content', utils.ContentFieldListFilter), ('likes',utils.LikesRangeFieldListFilter), ('create_time',utils.DateTimeRangeFilter),]
    search_fields = ['id','owner__name','owner__email','content']
    date_hierarchy = 'create_time'
    actions = [utils.export_to_excel_post]
    fieldsets = (
        (None,{'fields': ('owner',)}),
        (_('Post Content'),{'fields': ('content',)}),
        (_('Post Interactions'),{'fields': ('comments','likes','liked_users')}),
        (_('Imported Dates'),{'fields': ('create_time',)}),
    )
    readonly_fields = []

    add_fieldsets = (
        (None, {
            'classes': ('wide',), #Page look
            'fields': (
                'owner',
                'content',
                'like',
                'liked_users',
            ),
            }),
    
    )
class CommentsAdmin(admin.ModelAdmin):
    """Define the admin pages for comments."""
    ordering = ['id']
    list_display = ['id', 'owner', 'content', 'related_post', 'likes', 'create_time']
    list_filter = [('owner',utils.OwnerTextFieldListFilter), ('content', utils.ContentFieldListFilter), ('likes',utils.LikesRangeFieldListFilter),('create_time',utils.DateTimeRangeFilter)]
    search_fields = ['id','owner__name','owner__email','related_post__content','related_post__id']
    date_hierarchy = 'create_time'
    #date_hierarchy_drilldown = False
    actions = [utils.export_to_excel_comment]
    
    fieldsets = (
        (None,{'fields': ('owner',)}),
        (_('Comment Content'),{'fields': ('content',)}),
        (_('Comment Interactions'),{'fields': ('likes','liked_users')}),
        (_('Related Post'),{'fields': ('related_post',)}),
        (_('Imported Dates'),{'fields': ('create_time',)}),
    )
    readonly_fields = []
    add_fieldsets = (
        (None, {
            'classes': ('wide','extrapretty'), #Page look
            'fields': (
                'owner',
                'content',
                'like',
                'liked_users',
                'related_post',                
            ),
            }),
    )
admin.site.register(models.User, UserAdmin)
admin.site.register(models.Followship,FollowerAdmin)
admin.site.register(models.Comment,CommentsAdmin)
admin.site.register(models.Post,PostsAdmin)

